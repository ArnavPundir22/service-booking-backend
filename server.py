from flask import Flask, request, jsonify
from flask_cors import CORS
import uuid
import os
from openpyxl import Workbook, load_workbook
import requests

# -------------------------
# 1️⃣ Create Flask app
# -------------------------
app = Flask(__name__)
CORS(app)

# -------------------------
# 2️⃣ Send Email Function using Brevo API
# -------------------------
def send_email(recipient_email, subject, html_content):
    url = "https://api.brevo.com/v3/smtp/email"
    headers = {
        "accept": "application/json",
        "api-key": os.getenv("BREVO_API_KEY"),
        "content-type": "application/json"
    }
    data = {
        "sender": {"name": "SevaSetu Services", "email": "sevasetu.services@gmail.com"},
        "to": [{"email": recipient_email}],
        "subject": subject,
        "htmlContent": html_content
    }

    response = requests.post(url, headers=headers, json=data)
    print("📨 Email response:", response.status_code, response.text)
    return response.status_code in [200, 201]

# -------------------------
# 3️⃣ Route for booking
# -------------------------
@app.route("/send-booking", methods=["POST"])
def send_booking():
    data = request.json
    name = data.get("name")
    email = data.get("email")
    phone = data.get("phone")
    whatsapp = data.get("whatsapp")
    service = data.get("service")
    date = data.get("date")
    time = data.get("time")
    job = data.get("job", "Not specified")
    location = data.get("location", "")
    provider = data.get("provider", "Not assigned")

    # Generate a unique appointment ID
    appointment_id = str(uuid.uuid4()).split("-")[0].upper()

    # -------------------------
    # Email Content
    # -------------------------
    body = f"""
    <html>
        <body>
            <h2 style="color:#1E3A8A;">New Booking Received!</h2>
            <p><strong>Appointment ID:</strong> <span style="color:#DC2626;">{appointment_id}</span></p>
            <p><strong>Service Provider:</strong> {provider}</p>
            <p><strong>Customer Name:</strong> {name}</p>
            <p><strong>Email:</strong> {email}</p>
            <p><strong>Contact No.:</strong> {phone}</p>
            <p><strong>WhatsApp No.:</strong> {whatsapp}</p>
            <p><strong>Service:</strong> {service}</p>
            <p><strong>Job Description:</strong> {job}</p>
            <p><strong>Date & Time:</strong> {date} at {time}</p>
            <p><strong>Location:</strong> {location}</p>
            <hr/>
            <p style="color:gray;">You will be contacted shortly for service charges and payment proceeding...</p>
            <p style="color:gray;">This is an automated booking notification.</p>
            <p style="color:gray;">For any query, contact us at <b>sevasetu.services@gmail.com</b></p>
        </body>
    </html>
    """

    try:
        # -------------------------
        # Send Emails via Brevo
        # -------------------------
        send_email("sevasetu.services@gmail.com", f"New Booking: {service} by {name}", body)
        send_email(email, f"Your Appointment Confirmation (ID: {appointment_id})", body)

        # -------------------------
        # Save to Excel
        # -------------------------
        excel_file = "bookings.xlsx"
        if os.path.exists(excel_file):
            wb = load_workbook(excel_file)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append([
                "Appointment ID", "Provider", "Customer Name", "Email",
                "Contact No.", "WhatsApp No.", "Service", "Job",
                "Date", "Time", "Location"
            ])

        ws.append([
            appointment_id, provider, name, email,
            phone, whatsapp, service, job, date, time, location
        ])
        wb.save(excel_file)

        # -------------------------
        # Return success
        # -------------------------
        return jsonify({
            "success": True,
            "appointment_id": appointment_id,
            "message": f"✅ Booking Confirmed! Your Appointment ID is {appointment_id}"
        })

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

# -------------------------
# Run Flask App
# -------------------------
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
